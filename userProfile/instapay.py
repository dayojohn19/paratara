from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook


BANKS_SHEET_NAME = "Banks"
DETAILS_SHEET_NAME = "Details"
SELECT_BANK_PLACEHOLDERS = {"Select Bank...", "Select Bank…"}
DETAILS_HEADERS = (
    "Bank Name",
    "Bank Account Name",
    "Bank Account Number",
    "Amount",
    "Remarks",
)


def get_instapay_template_path():
    return Path(__file__).resolve().parent / "instapay_template.xlsx"


@lru_cache(maxsize=1)
def get_instapay_bank_names():
    template_path = get_instapay_template_path()
    if not template_path.exists():
        return tuple()

    workbook = load_workbook(template_path, read_only=True, data_only=True)
    try:
        if BANKS_SHEET_NAME not in workbook.sheetnames:
            return tuple()

        bank_names = []
        seen_names = set()
        banks_sheet = workbook[BANKS_SHEET_NAME]
        for row in banks_sheet.iter_rows(min_col=1, max_col=1, values_only=True):
            value = str(row[0]).strip() if row and row[0] is not None else ""
            if not value or value in seen_names:
                continue
            bank_names.append(value)
            seen_names.add(value)

        return tuple(bank_names)
    finally:
        workbook.close()


def get_instapay_bank_choices():
    return tuple((bank_name, bank_name) for bank_name in get_instapay_bank_names())


INSTAPAY_BANK_CHOICES = get_instapay_bank_choices()


def _clean_cell_value(value):
    return str(value).strip() if value is not None else ""


def _find_details_headers(details_sheet):
    max_row = max(details_sheet.max_row, 1)
    max_column = max(details_sheet.max_column, len(DETAILS_HEADERS))
    required_headers = set(DETAILS_HEADERS)

    for row_number in range(1, max_row + 1):
        header_map = {}
        for column_number in range(1, max_column + 1):
            header_value = _clean_cell_value(details_sheet.cell(row=row_number, column=column_number).value)
            if header_value in required_headers and header_value not in header_map:
                header_map[header_value] = column_number

        if required_headers.issubset(header_map):
            return row_number, header_map

    expected_headers = ", ".join(DETAILS_HEADERS)
    raise ValueError(f"Could not find the Details header row with: {expected_headers}.")


def _is_available_details_row(details_sheet, row_number, header_map):
    values = [
        _clean_cell_value(details_sheet.cell(row=row_number, column=header_map[header]).value)
        for header in DETAILS_HEADERS
    ]
    if values[0] in SELECT_BANK_PLACEHOLDERS and not any(values[1:]):
        return True
    return not any(values)


def _next_details_row(details_sheet, header_row, header_map):
    max_row = max(details_sheet.max_row, header_row + 1)
    for row_number in range(header_row + 1, max_row + 1):
        if _is_available_details_row(details_sheet, row_number, header_map):
            return row_number
    return max_row + 1


def save_instapay_template(*, bank_name, bank_account_name, bank_account_number, amount=None, remarks=None):
    template_path = get_instapay_template_path()
    if not template_path.exists():
        raise FileNotFoundError(f"InstaPay template not found: {template_path}")

    workbook = load_workbook(template_path)
    try:
        if DETAILS_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Could not find the {DETAILS_SHEET_NAME!r} sheet in the InstaPay template.")

        details_sheet = workbook[DETAILS_SHEET_NAME]
        header_row, header_map = _find_details_headers(details_sheet)
        row_number = _next_details_row(details_sheet, header_row, header_map)
        details_sheet.cell(row=row_number, column=header_map["Bank Name"]).value = bank_name
        details_sheet.cell(row=row_number, column=header_map["Bank Account Name"]).value = bank_account_name
        details_sheet.cell(row=row_number, column=header_map["Bank Account Number"]).value = bank_account_number
        details_sheet.cell(row=row_number, column=header_map["Amount"]).value = amount
        details_sheet.cell(row=row_number, column=header_map["Remarks"]).value = remarks
        workbook.save(template_path)
        return row_number
    finally:
        workbook.close()
